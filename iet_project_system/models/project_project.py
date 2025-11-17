from odoo import models, fields, api
from datetime import timedelta
import xlsxwriter
import io
from io import BytesIO
import base64
import openpyxl
import logging

_logger = logging.getLogger(__name__)


class Project(models.Model):
    _inherit = 'project.project'

    project_plan_line_ids = fields.One2many(
        'project.plan.line',
        'project_id',
        string='Project Plan Lines'
    )
    hide_button = fields.Boolean(string="checkbox")

    scope_ids = fields.Many2many('project.scope', string='Scope')
    completion_percent = fields.Float(string='Completion %', compute='_compute_completion_percent', store=True)

    def _cron_send_deadline_notifications(self):
        """إرسال تذكيرات قبل انتهاء المواعيد (0-10 أيام)"""
        today = fields.Date.today()

        # المستخدمين المحددين + مدير المشروع
        fixed_users = self.env['res.users'].search([
            ('name', 'in', ['Omar elnabawy', 'Mahmoud Elaskary', 'Shrouq Abdeldaym'])
        ])

        # جلب كل المشاريع
        projects = self.search([])
        if not projects:
            _logger.info("لا توجد مشاريع للتحقق من المواعيد.")
            return

        # نوع الأكتيفيتي (To Do)
        activity_type = self.env.ref('mail.mail_activity_data_todo', raise_if_not_found=False)
        if not activity_type:
            activity_type = self.env['mail.activity.type'].search([('name', '=', 'To Do')], limit=1)

        for project in projects:
            # تعريف المواعيد المراد متابعتها
            dates = {
                'Project End': project.end_project_date,
                'Free Support End': project.free_support_end_date,
                'Contract Support End': project.contract_project_end_date
            }

            for label, date in dates.items():
                if not date:
                    continue

                diff = (date - today).days
                _logger.info("Project: %s | %s ends in %d days", project.name, label, diff)

                # فقط إذا كان الموعد خلال 0 إلى 10 أيام
                if 0 <= diff <= 10:
                    recipients = fixed_users | project.user_id
                    body = f"Reminder: {label} for project '<strong>{project.name}</strong>' ends on <strong>{date}</strong>."

                    for user in recipients:
                        if not user.partner_id:
                            _logger.warning("User %s has no partner_id!", user.name)
                            continue

                        partner_id = user.partner_id.id
                        _logger.info("Sending notifications to %s (%s)", user.name, user.email or 'No email')

                        # 1. رسالة في الـ Chatter
                        project.message_post(
                            body=body,
                            partner_ids=[partner_id],
                            subtype_xmlid='mail.mt_comment'
                        )

                        # 2. إرسال إيميل (إن وجد)
                        if user.email:
                            try:
                                mail_values = {
                                    'subject': f"Deadline Reminder: {project.name} - {label}",
                                    'body_html': f"<p>{body}</p><p>Please check the project for details.</p>",
                                    'email_to': user.email,
                                    'email_from': self.env.user.email or 'no-reply@iet.com',
                                }
                                mail = self.env['mail.mail'].create(mail_values)
                                mail.send(auto_commit=True)
                                _logger.info("Email sent to %s", user.email)
                            except Exception as e:
                                _logger.error("Failed to send email to %s: %s", user.email, str(e))

                        # 3. إنشاء أكتيفيتي (تظهر في My Activities)
                        try:
                            self.env['mail.activity'].create({
                                'res_id': project.id,
                                'res_model_id': self.env['ir.model']._get_id('project.project'),
                                'activity_type_id': activity_type.id if activity_type else False,
                                'summary': f"{label} ends in {diff} day(s)",
                                'note': f"<p>{body}</p>",
                                'date_deadline': today,  # أو date - 1
                                'user_id': user.id,
                            })
                            _logger.info("Activity created for %s", user.name)
                        except Exception as e:
                            _logger.error("Failed to create activity for %s: %s", user.name, str(e))

                        # 4. نوتفيكيشن فوري في الجرس (Bell Icon) - Odoo 18
                        try:
                            project.message_notify(
                                partner_ids=[partner_id],
                                body=body,
                                subject=f"{label} Alert: {project.name}",
                            )
                            _logger.info("Bell notification sent to %s", user.name)
                        except Exception as e:
                            _logger.error("Failed to send notification to %s: %s", user.name, str(e))

    @api.depends('project_plan_line_ids.status_done')
    def _compute_completion_percent(self):
        for project in self:
            total = len(project.project_plan_line_ids)
            done = len(project.project_plan_line_ids.filtered(lambda l: l.status_done))
            project.completion_percent = (done / total) * 100 if total else 0

    def action_generate_tasks(self):
        Task = self.env['project.task']
        for project in self:
            project.hide_button = True
            project.project_plan_line_ids.assign_milestones_to_plan_lines()
            task_type_ids = self.env['project.task.type'].search([
                ('generate_tasks', '=', True),
                ('project_ids', 'in', [project.id])
            ])
            if not task_type_ids:
                _logger.warning("No task types found for project %s with generate_tasks=True", project.name)
                continue

            for plan_line in project.project_plan_line_ids.filtered(lambda l: not l.display_type):
                vals = {
                    'name': plan_line.name,
                    'project_id': project.id,
                    'date_start': plan_line.planned_start_date,
                    'end_date': plan_line.planned_end_date,
                    'team_name': project.team_id.name,
                    'user_ids': [(6, 0, [project.user_id.id])] if project.user_id else False,
                    'milestone_id': plan_line.milestone_id.id if plan_line.milestone_id else False,
                    'stage_id': task_type_ids[0].id,
                }
                _logger.info("Creating task with vals: %s", vals)
                Task.create(vals)

    def action_update_tasks(self):
        Task = self.env['project.task']
        for project in self:
            project.hide_button = True
            project.project_plan_line_ids.assign_milestones_to_plan_lines()
            task_type_ids = self.env['project.task.type'].search([
                ('generate_tasks', '=', True),
                ('project_ids', 'in', [project.id])
            ])
            if not task_type_ids:
                _logger.warning("No task types found for project %s with generate_tasks=True", project.name)
                continue

            for plan_line in project.project_plan_line_ids.filtered(lambda l: not l.display_type):
                vals = {
                    'name': plan_line.name,
                    'project_id': project.id,
                    'date_start': plan_line.planned_start_date,
                    'end_date': plan_line.planned_end_date,
                    'team_name': project.team_id.name,
                    'user_ids': [(6, 0, [project.user_id.id])] if project.user_id else False,
                    'milestone_id': plan_line.milestone_id.id if plan_line.milestone_id else False,
                    'stage_id': task_type_ids[0].id,

                }

                if plan_line.task_id:
                    _logger.info("Updating task %s with vals: %s", plan_line.task_id.id, vals)
                    plan_line.task_id.with_context(skip_stage_validation=True).write(vals)
                else:
                    _logger.info("Creating new task with vals: %s", vals)
                    task = Task.with_context(skip_stage_validation=True).create(vals)
                    plan_line.task_id = task

    def action_print_project_plan(self):
        for project in self:
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            worksheet = workbook.add_worksheet("Project Plan")

            header_format = workbook.add_format({'bold': True, 'bg_color': '#CFE2F3', 'border': 1})
            task_format = workbook.add_format({'border': 1})
            title_format = workbook.add_format({'bold': True, 'font_size': 14})
            date_format = workbook.add_format({'italic': True, 'align': 'right'})

            worksheet.set_column('A:H', 22)

            logo = project.company_id.logo
            if logo:
                image_data = io.BytesIO(base64.b64decode(logo))
                worksheet.insert_image('A1', 'logo.png', {
                    'image_data': image_data,
                    'x_scale': 0.5,
                    'y_scale': 0.5,
                    'x_offset': 0,
                    'y_offset': 0,
                    'object_position': 1
                })

            worksheet.merge_range('C1:D1', project.company_id.name or '', title_format)

            current_date = fields.Date.context_today(project)
            worksheet.merge_range('E1:F1', f'Date: {current_date.strftime("%Y-%m-%d")}', date_format)

            headers = ["Task Name", "Planned Start Date", "Actual Start Date",
                       "Planned End Date", "Actual End Date", "Task Owner", "Done", "Comments"]
            header_row = 3
            for col, header in enumerate(headers):
                worksheet.write(header_row, col, header, header_format)

            row = header_row + 1

            plan_lines = project.project_plan_line_ids
            for line in plan_lines:
                worksheet.write(row, 0, line.name or '', task_format)
                worksheet.write(row, 1, line.planned_start_date.strftime(
                    "%Y-%m-%d %H:%M") if line.planned_start_date else '', task_format)
                worksheet.write(row, 2, line.actual_start_date.strftime(
                    "%Y-%m-%d %H:%M") if line.actual_start_date else '', task_format)
                worksheet.write(row, 3, line.planned_end_date.strftime(
                    "%Y-%m-%d %H:%M") if line.planned_end_date else '', task_format)
                worksheet.write(row, 4, line.actual_end_date.strftime(
                    "%Y-%m-%d %H:%M") if line.actual_end_date else '', task_format)
                worksheet.write(row, 5, line.task_owner or '', task_format)
                worksheet.write(row, 6, line.status_done or '', task_format)
                worksheet.write(row, 7, line.comments or '', task_format)
                row += 1

            workbook.close()
            output.seek(0)

            attachment = self.env['ir.attachment'].create({
                'name': f'{project.name}_project_plan.xlsx',
                'type': 'binary',
                'datas': base64.b64encode(output.read()),
                'res_model': self._name,
                'res_id': project.id,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            })

            download_url = f'/web/content/{attachment.id}?download=true'
            return {
                'type': 'ir.actions.act_url',
                'url': download_url,
                'target': 'self',
            }

    def action_import_plan(self):
        """استيراد خطة المشروع من ملف Excel (xlsx) باستخدام openpyxl"""
        if not self.excel_file:
            raise UserWarning("Please upload an Excel file first!")

        try:
            # فك تشفير الملف
            file_content = base64.b64decode(self.excel_file)
            file_stream = BytesIO(file_content)

            # قراءة ملف Excel باستخدام openpyxl
            workbook = openpyxl.load_workbook(file_stream, data_only=True)
            sheet = workbook.active

            # إيجاد صف الهيدر (يبدأ بـ Task Name)
            header_row = None
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False):
                if row[0].value and 'Task Name' in str(row[0].value):
                    header_row = row[0].row
                    break

            if not header_row:
                raise UserWarning("Cannot find header row with 'Task Name'")

            plan_lines_to_create = []

            # قراءة البيانات بعد الهيدر
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                task_name = row[0]
                if not task_name or str(task_name).strip() == '':
                    continue

                # قراءة التواريخ
                planned_start = row[1]
                actual_start = row[2]
                planned_end = row[3]
                actual_end = row[4]

                task_owner = row[5] if len(row) > 5 else ''
                done = row[6] if len(row) > 6 else ''
                comments = row[7] if len(row) > 7 else ''

                # تحويل Done إلى boolean
                status_done = False
                if done:
                    done_str = str(done).strip().lower()
                    status_done = done_str in ['true', '1', 'yes', 'done', 'x']

                vals = {
                    'project_id': self.project_id.id,
                    'name': str(task_name).strip(),
                    'planned_start_date': planned_start,
                    'actual_start_date': actual_start,
                    'planned_end_date': planned_end,
                    'actual_end_date': actual_end,
                    'task_owner': str(task_owner).strip() if task_owner else '',
                    'status_done': status_done,
                    'comments': str(comments).strip() if comments else '',
                }

                plan_lines_to_create.append(vals)

            # إنشاء الخطوط
            if plan_lines_to_create:
                self.env['project.plan.line'].create(plan_lines_to_create)

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Success',
                    'message': f'Successfully imported {len(plan_lines_to_create)} tasks',
                    'type': 'success',
                    'sticky': False,
                },
            }

        except Exception as e:
            _logger.error(f"Error importing project plan: {str(e)}")
            raise UserWarning(f"Error importing file: {str(e)}")

    def action_import_project_plan(self):
        action = self.env["ir.actions.actions"]._for_xml_id("iet_project_system.action_import_project_plan_wizard")

        return action
