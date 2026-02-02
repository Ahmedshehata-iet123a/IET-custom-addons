from odoo import models, fields, api
from datetime import datetime
import base64
import io
import logging
import openpyxl

_logger = logging.getLogger(__name__)


class ProjectImportPlan(models.TransientModel):
    _name = 'project.import.plan'
    _description = 'Import Project Plan from Excel'

    excel_file = fields.Binary(string='Excel File', required=True)
    filename = fields.Char(string='Filename')
    project_id = fields.Many2one(
        'project.project',
        string='Project',
        required=True,
        default=lambda self: self.env.context.get('default_project_id')
    )

    def action_import_plan(self):
        """استيراد خطة المشروع من ملف Excel (xlsx فقط)"""
        if not self.excel_file:
            raise UserWarning("Please upload an Excel file first!")

        try:
            file_content = base64.b64decode(self.excel_file)
            workbook = openpyxl.load_workbook(
                filename=io.BytesIO(file_content),
                data_only=True
            )
            sheet = workbook.active

            # البحث عن صف الهيدر
            header_row = None
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if row and row[0] and 'Task Name' in str(row[0]):
                    header_row = row_idx
                    break

            if header_row is None:
                raise UserWarning("Cannot find header row with 'Task Name'")

            plan_lines_to_create = []
            rows = list(sheet.iter_rows(values_only=True))[header_row + 1:]

            for row_idx, row in enumerate(rows, start=header_row + 1):
                try:
                    task_name = row[0]
                    if not task_name or str(task_name).strip() == '':
                        continue

                    # ✅ NEW FIELDS
                    milestone_type_new = row[1] if len(row) > 1 else False
                    milestone_weight = row[2] if len(row) > 2 else False

                    # التواريخ
                    planned_start = self._parse_date(row[3]) if len(row) > 3 else False
                    actual_start = self._parse_date(row[4]) if len(row) > 4 else False
                    planned_end = self._parse_date(row[5]) if len(row) > 5 else False
                    actual_end = self._parse_date(row[6]) if len(row) > 6 else False

                    task_owner = row[7] if len(row) > 7 else ''
                    done = row[8] if len(row) > 8 else ''
                    comments = row[9] if len(row) > 9 else ''

                    status_done = str(done).strip().lower() in ['true', '1', 'yes', 'done', 'x']

                    if milestone_type_new:
                        milestone_type_rec = self.env['milestone.type'].search([('name', '=', str(milestone_type_new).strip())], limit=1)
                        if not milestone_type_rec:
                            milestone_type_rec = self.env['milestone.type'].create({'name': str(milestone_type_new).strip()})
                        milestone_type_id = milestone_type_rec.id
                    else:
                        milestone_type_id = False

                    try:
                        weight = int(milestone_weight) if milestone_weight else 0
                    except (ValueError, TypeError):
                        weight = 0

                    vals = {
                        'project_id': self.project_id.id,
                        'name': str(task_name).strip(),

                        'milestone_type_new': milestone_type_id,
                        'milestone_weight': weight,
                        'display_type': 'line_section' if milestone_type_id else False,

                        'planned_start_date': planned_start,
                        'actual_start_date': actual_start,
                        'planned_end_date': planned_end,
                        'actual_end_date': actual_end,
                        'task_owner': str(task_owner).strip() if task_owner else '',
                        'status_done': status_done,
                        'comments': str(comments).strip() if comments else '',
                    }

                    plan_lines_to_create.append(vals)

                except Exception as e:
                    _logger.warning(f"Error processing row {row_idx}: {str(e)}")
                    continue

            if plan_lines_to_create:
                self.env['project.plan.line'].create(plan_lines_to_create)
                _logger.info(
                    f"Successfully imported {len(plan_lines_to_create)} plan lines"
                )

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Success',
                    'message': f'Successfully imported {len(plan_lines_to_create)} tasks',
                    'type': 'success',
                    'sticky': False,
                    'next': {'type': 'ir.actions.act_window_close'},
                }
            }

        except Exception as e:
            _logger.error(f"Error importing project plan: {str(e)}")
            raise UserWarning(f"Error importing file: {str(e)}")

    def _parse_date(self, value):
        if not value:
            return False

        if isinstance(value, datetime):
            return value

        date_str = str(value).strip()
        for fmt in [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y-%m-%d',
            '%d/%m/%Y',
            '%m/%d/%Y',
            '%d-%m-%Y'
        ]:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue

        _logger.warning(f"Cannot parse date: {value}")
        return False
